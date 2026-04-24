###################################################################################################;[-]
# Filename: gen_xml_to_excel.py
# Purpose : This script parses XML register description files and generates a structured Excel
#           register map with formatted styling. It supports multi-sheet generation, applies
#           color coding for different sheets, and highlights reserved/unused fields for
#           better readability and documentation.
# Features:
# - Converts XML register data into Excel format
# - Supports multiple XML inputs → multiple Excel sheets
# - Applies sheet-specific colors for easy differentiation
# - Highlights 'reserved' / 'unused' fields in red
# - Auto-adjusts column width and applies filters
# - Generates professional, documentation-ready output
# Version : 1.0
# Author  : Susmitha
# Date    : 24/04/2026
###################################################################################################

import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

def xml_to_excel_final(tasks, final_excel_name):
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # --- COMMON STYLES ---
    header_fill = PatternFill(
        start_color="FFD9D9D9",
        end_color="FFD9D9D9",
        fill_type="solid"
    )
    header_font = Font(color="000000", bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Red fill (same for all sheets)
    RED_FILL = PatternFill(
        start_color="FFFF9999",
        end_color="FFFF9999",
        fill_type="solid"
    )

    for xml_file, sheet_name, sheet_color in tasks:

        if not os.path.exists(xml_file):
            print(f"Skipping: {xml_file} not found.")
            continue

        # 🎨 Convert sheet color to ARGB
        base_color = "FF" + sheet_color.upper()

        BASE_FILL = PatternFill(
            start_color=base_color,
            end_color=base_color,
            fill_type="solid"
        )

        tree = ET.parse(xml_file)
        root = tree.getroot()

        ws = wb.create_sheet(title=sheet_name)

        all_rows = []
        headers = []

        # --- EXTRACT DATA ---
        for reg_node in root.findall('.//Register'):

            if len(reg_node) > 0:
                reg_info = {}

                for child in reg_node:
                    if len(child) == 0 and child.text:
                        reg_info[child.tag] = child.text.strip()

                fields = reg_node.findall('Field')

                if fields:
                    for f_node in fields:
                        row_data = reg_info.copy()

                        for f_child in f_node:
                            if f_child.text:
                                row_data[f_child.tag] = f_child.text.strip()

                        all_rows.append(row_data)
                else:
                    all_rows.append(reg_info)

        # --- HEADERS ---
        for row in all_rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)

        # --- WRITE HEADER ---
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        # --- WRITE DATA ---
        for row_idx, row_data in enumerate(all_rows, 2):

            # Detect field safely
            field_value = ""
            for key in row_data.keys():
                if key.lower() in ["field", "name"]:
                    field_value = str(row_data.get(key, "")).lower()
                    break

            # 🎯 COLOR LOGIC
            if "unused" in field_value or "reserved" in field_value:
                fill_to_apply = RED_FILL
            else:
                fill_to_apply = BASE_FILL

            for col_idx, header in enumerate(headers, 1):
                val = row_data.get(header, "")

                cell = ws.cell(row=row_idx, column=col_idx, value=val)

                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = fill_to_apply  # Apply color

        # --- AUTO WIDTH ---
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter

            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_len + 3

        ws.auto_filter.ref = ws.dimensions

    wb.save(final_excel_name)
    print(f"✅ SUCCESS: '{final_excel_name}' generated with multi-sheet colors.")


# ==========================
# RUN
# ==========================
if __name__ == "__main__":

    my_tasks = [
        ('xml1.xml', 'ADC_Registers', 'FFFFCC'),   # Light Yellow
        ('xml2.xml', 'DLL_Registers', 'CCFFCC'),   # Light Green
        ('xml3.xml', 'Slave_Config', 'CCE5FF')     # Light Blue
    ]

    xml_to_excel_final(my_tasks, 'register_map_colored.xlsx')